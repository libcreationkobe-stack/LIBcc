# -*- coding: utf-8 -*-
"""Instagram採用KPI管理シートを生成する。

数値（青ヘッダーの入力欄）を入れるだけで、各ファネルの率が自動で出るように
数式・書式込みの .xlsx を出力する。Googleスプレッドシートに変換して使う想定。
"""

import argparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

MONTHS = ["8月", "9月", "10月", "11月", "12月", "1月", "2月",
          "3月", "4月", "5月", "6月", "7月", "8月"]

HEADER_ROW = 2
GUIDE_ROW = 3                               # 悪い／普通／良いの目安
FIRST_ROW = 4
LAST_ROW = FIRST_ROW + len(MONTHS) - 1      # 16
TOTAL_ROW = LAST_ROW + 1                    # 17
AVG_ROW = LAST_ROW + 2                      # 18

# 悪い／普通／良いの判定ライン。bad未満＝赤、bad〜good＝緑、good以上＝黄色。
BENCHMARKS = {
    "O": (0.03, 0.05, "悪い 〜3% / 普通 3〜5% / 良い 5%〜"),
    "P": (0.05, 0.10, "悪い 〜5% / 普通 5〜10% / 良い 10%〜"),
    "Q": (0.20, 0.40, "悪い 〜20% / 普通 20〜40% / 良い 40%〜"),
    "V": (0.00005, 0.0002, "悪い 〜0.005% / 普通 0.005〜0.02% / 良い 0.02%〜"),
}

# (列, ヘッダー, 種別, 数式テンプレート, 表示形式, 幅)
#   種別 input = 手入力 / calc = 自動計算
COLUMNS = [
    ("A", "月",                  "input", None, "@",        9),
    ("B", "投稿数",              "input", None, "#,##0",    9),
    ("C", "インプレッション",    "input", None, "#,##0",   15),
    ("D", "リーチ数",            "input", None, "#,##0",   12),
    ("E", "プロフアクセス数",    "input", None, "#,##0",   15),
    ("F", "リンクタップ数",      "input", None, "#,##0",   14),
    ("G", "LINE登録",            "input", None, "#,##0",   11),
    ("H", "面接",                "input", None, "#,##0",    9),
    ("I", "採用数(LINE経由)",    "input", None, "#,##0",   15),
    ("J", "その他問い合わせ",    "input", None, "#,##0",   15),
    ("K", "採用数(その他経由)",  "input", None, "#,##0",   16),
    ("L", "採用数 合計",         "calc",  '=IFERROR(IF(COUNT(I{r}:K{r})=0,"",SUM(I{r},K{r})),"")', "#,##0", 11),
    ("M", "1投稿あたり\nインプ", "calc",  '=IFERROR(C{r}/B{r},"")',        "#,##0",  13),
    ("N", "リーチ率\n(リーチ/インプ)",           "calc", '=IFERROR(D{r}/C{r},"")', "0.0%", 13),
    ("O", "プロフアクセス率\n(プロフ/リーチ)",   "calc", '=IFERROR(E{r}/D{r},"")', "0.0%", 15),
    ("P", "リンクタップ率\n(タップ/プロフ)",     "calc", '=IFERROR(F{r}/E{r},"")', "0.0%", 15),
    ("Q", "LINE登録率\n(登録/タップ)",           "calc", '=IFERROR(G{r}/F{r},"")', "0.0%", 14),
    ("R", "面接率\n(面接/LINE登録)",             "calc", '=IFERROR(H{r}/G{r},"")', "0.0%", 14),
    ("S", "採用率\n(採用/面接)",                 "calc", '=IFERROR(I{r}/H{r},"")', "0.0%", 13),
    ("T", "LINE→採用率\n(採用/LINE登録)",        "calc", '=IFERROR(I{r}/G{r},"")', "0.0%", 15),
    ("U", "その他採用率\n(採用/その他問合せ)",   "calc", '=IFERROR(K{r}/J{r},"")', "0.0%", 16),
    ("V", "リーチ→採用率\n(採用合計/リーチ)",    "calc", '=IFERROR(L{r}/D{r},"")', "0.000%", 15),
    ("W", "1採用あたり\n投稿数",                 "calc", '=IFERROR(B{r}/L{r},"")', "#,##0.0", 13),
]

BAD_FILL = PatternFill("solid", fgColor="F4CCCC")
OK_FILL = PatternFill("solid", fgColor="D9EAD3")
GOOD_FILL = PatternFill("solid", fgColor="FFF2CC")
BAD_FONT = Font(color="990000", bold=True, size=10)
OK_FONT = Font(color="274E13", bold=True, size=10)
GOOD_FONT = Font(color="7F6000", bold=True, size=10)

TITLE_FILL = PatternFill("solid", fgColor="1F3864")
INPUT_FILL = PatternFill("solid", fgColor="2E75B6")
CALC_FILL = PatternFill("solid", fgColor="548235")
MONTH_FILL = PatternFill("solid", fgColor="F2F2F2")
CALC_CELL_FILL = PatternFill("solid", fgColor="EDF3E9")
SUM_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
MED = Side(style="medium", color="808080")


def build(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "月次KPI"

    # --- タイトル行 -------------------------------------------------
    ws["A1"] = "Instagram採用 月次KPI（青＝入力欄／緑＝自動計算）"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    ws.row_dimensions[1].height = 24

    # --- ヘッダー ---------------------------------------------------
    for col, header, kind, _f, _fmt, width in COLUMNS:
        cell = ws[f"{col}{HEADER_ROW}"]
        cell.value = header
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = INPUT_FILL if kind == "input" else CALC_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=MED)
        ws.column_dimensions[col].width = width
    ws.row_dimensions[HEADER_ROW].height = 42

    # --- 目安行（悪い／普通／良い）----------------------------------
    ws[f"A{GUIDE_ROW}"] = "目安"
    ws[f"A{GUIDE_ROW}"].font = Font(bold=True, size=9)
    ws[f"A{GUIDE_ROW}"].fill = MONTH_FILL
    ws[f"A{GUIDE_ROW}"].alignment = Alignment(horizontal="center", vertical="center")
    for col, _h, _k, _f, _fmt, _w in COLUMNS[1:]:
        cell = ws[f"{col}{GUIDE_ROW}"]
        if col in BENCHMARKS:
            cell.value = BENCHMARKS[col][2]
        cell.font = Font(size=8)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ws.row_dimensions[GUIDE_ROW].height = 30

    # --- 明細行 -----------------------------------------------------
    for i, month in enumerate(MONTHS):
        r = FIRST_ROW + i
        for col, _h, kind, formula, fmt, _w in COLUMNS:
            cell = ws[f"{col}{r}"]
            if col == "A":
                cell.value = month
                cell.fill = MONTH_FILL
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif kind == "calc":
                cell.value = formula.format(r=r)
                cell.fill = CALC_CELL_FILL
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = fmt
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        ws.row_dimensions[r].height = 20

    # --- 合計行 / 平均行 --------------------------------------------
    for label, row in (("合計", TOTAL_ROW), ("平均", AVG_ROW)):
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, size=10)
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        for col, _h, kind, formula, fmt, _w in COLUMNS:
            if col == "A":
                cell = ws[f"A{row}"]
            else:
                cell = ws[f"{col}{row}"]
                rng = f"{col}{FIRST_ROW}:{col}{LAST_ROW}"
                if kind == "input" or col == "L":
                    # 実数列は合計。平均行は月平均。
                    if row == TOTAL_ROW:
                        cell.value = f'=IFERROR(IF(COUNT({rng})=0,"",SUM({rng})),"")'
                    else:
                        cell.value = f'=IFERROR(IF(COUNT({rng})=0,"",AVERAGE({rng})),"")'
                    cell.number_format = "#,##0" if row == TOTAL_ROW else "#,##0.0"
                else:
                    # 率列は「合計値どうしで割り直した通期率」と「単純平均」
                    if row == TOTAL_ROW:
                        cell.value = formula.format(r=TOTAL_ROW)
                    else:
                        cell.value = f'=IFERROR(IF(COUNT({rng})=0,"",AVERAGE({rng})),"")'
                    cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.fill = SUM_FILL
            cell.font = Font(bold=True, size=10)
            cell.border = Border(left=THIN, right=THIN,
                                 top=MED if row == TOTAL_ROW else THIN, bottom=THIN)
        ws.row_dimensions[row].height = 20

    # --- 4つの率を 悪い＝赤 / 普通＝緑 / 良い＝黄色 で色分け ---------
    for col, (bad, good, _label) in BENCHMARKS.items():
        rng = f"{col}{FIRST_ROW}:{col}{AVG_ROW}"
        top = f"${col}{FIRST_ROW}"
        # 空欄を赤くしないよう ISNUMBER で必ず絞る。
        tiers = [
            (f"AND(ISNUMBER({top}),{top}<{bad})", BAD_FILL, BAD_FONT),
            (f"AND(ISNUMBER({top}),{top}>={bad},{top}<{good})", OK_FILL, OK_FONT),
            (f"AND(ISNUMBER({top}),{top}>={good})", GOOD_FILL, GOOD_FONT),
        ]
        for formula, fill, font in tiers:
            ws.conditional_formatting.add(
                rng, FormulaRule(formula=[formula], fill=fill, font=font))

    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = True
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(COLUMNS))}{HEADER_ROW}"

    # --- 指標の定義シート -------------------------------------------
    ws2 = wb.create_sheet("指標の定義")
    ws2["A1"] = "指標の定義（率はすべて自動計算。入力欄が空なら空欄のまま）"
    ws2["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws2["A1"].fill = TITLE_FILL
    ws2.merge_cells("A1:C1")
    ws2.row_dimensions[1].height = 24

    defs = [
        ("列", "指標", "計算式・意味"),
        ("B〜K", "入力欄", "毎月の実数を手で入れる（インサイト・LINE・面接記録から転記）"),
        ("L", "採用数 合計", "採用数(LINE経由) ＋ 採用数(その他経由)"),
        ("M", "1投稿あたりインプ", "インプレッション ÷ 投稿数。投稿の平均パワー"),
        ("N", "リーチ率", "リーチ数 ÷ インプレッション。低いと同じ人に何度も出ている"),
        ("O", "プロフアクセス率", "プロフアクセス数 ÷ リーチ数。投稿→プロフの興味喚起力"),
        ("P", "リンクタップ率", "リンクタップ数 ÷ プロフアクセス数。プロフ文とハイライトの出来"),
        ("Q", "LINE登録率", "LINE登録 ÷ リンクタップ数。LP・登録導線の出来"),
        ("R", "面接率", "面接 ÷ LINE登録。LINE内トークの出来"),
        ("S", "採用率", "採用数(LINE経由) ÷ 面接。面接の見極め・訴求力"),
        ("T", "LINE→採用率", "採用数(LINE経由) ÷ LINE登録。LINE1件の価値"),
        ("U", "その他採用率", "採用数(その他経由) ÷ その他問い合わせ"),
        ("V", "リーチ→採用率", "採用数 合計 ÷ リーチ数。全体の最終CVR"),
        ("W", "1採用あたり投稿数", "投稿数 ÷ 採用数 合計。1人採るのに必要な投稿本数"),
        ("", "", ""),
        ("3行", "目安", "4つの率に 悪い／普通／良い のラインを記載"),
        ("17行", "合計", "実数は年間合計、率は「合計値どうしを割り直した通期の率」"),
        ("18行", "平均", "実数は月平均、率は各月の率の単純平均"),
        ("", "", ""),
        ("色分け", "赤＝悪い / 緑＝普通 / 黄＝良い", "プロフアクセス率・リンクタップ率・LINE登録率・リーチ→採用率の4列に自動で色が付く"),
        ("目安", "プロフアクセス率 3〜5%", "Instagram運用の一般的な目標値。3%未満は投稿がプロフィールまで引っ張れていない"),
        ("目安", "リンクタップ率 5〜10%", "10%が分岐点。10%超ならプロフィール文とハイライトが機能している"),
        ("目安", "LINE登録率 20〜40%", "LINE友だち追加のCVR一般値。特典と導線が強いと50%超も出る"),
        ("目安", "リーチ→採用率 0.005〜0.02%", "公開ベンチマークがないため上の3つ＋面接率・採用率から逆算した値"),
    ]
    for i, row in enumerate(defs, start=3):
        for j, v in enumerate(row, start=1):
            c = ws2.cell(row=i, column=j, value=v)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if i == 3:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = CALC_FILL
            elif j == 2:
                c.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 62

    wb.save(path)
    print(f"saved: {path}")


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("-o", "--out", default="Instagram採用KPI管理シート.xlsx")
    build(p.parse_args().out)
